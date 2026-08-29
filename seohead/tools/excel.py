"""Excel / CSV I/O for the SEO toolkit (pure core, no GUI).

Faithful Python port of ``src/core/excel.ts``. Reads redirect workbooks
(xlsx / csv / tsv / txt) into old->new URL pairs and serialises keyword
clusters to CSV / XLSX.

Public API (matches the shared handler layer):
    parse_redirects_workbook(path) -> list[dict] | dict
    build_clusters_csv(clusters) -> str
    write_clusters_xlsx(clusters, path) -> None

A cluster is a mapping like ``{"label": str, "keywords": [str, ...]}``.
Optional keys ``id`` and ``count`` are honoured when present, otherwise
they are derived (1-based index and ``len(keywords)`` respectively).

Errors are returned as ``{"ok": False, "error": "..."}`` instead of raising,
so the caller process never crashes.
"""

from __future__ import annotations

import csv
import re
from typing import Any

try:  # openpyxl is an allowed dependency; guard so the module still imports.
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - only when openpyxl is missing
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]

# Header words that flag the first row of a redirects file as a header row.
REDIRECT_HEADER_RE = re.compile(r"^(old|new|url|старый|новый|from|to|откуда|куда)")

# Characters Excel forbids in a sheet name: \ / * ? : [ ]
_FORBIDDEN_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

_XLSX_EXTS = (".xlsx", ".xlsm", ".xltx", ".xltm")


# ── redirects ────────────────────────────────────────────────────────────────
def _read_rows(path: str) -> list[list[Any]]:
    """Read a spreadsheet/delimited file into a list of non-blank rows.

    Missing cells are normalised to '' and fully blank rows are dropped, to
    mirror xlsx ``sheet_to_json({ header: 1, defval: '', blankrows: false })``.
    """
    lower = path.lower()
    rows: list[list[Any]] = []

    if lower.endswith(_XLSX_EXTS):
        if load_workbook is None:
            raise RuntimeError("openpyxl is not available")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if ws is None:
            wb.close()
            raise RuntimeError("no sheet")
        for raw in ws.iter_rows(values_only=True):
            cells = ["" if c is None else c for c in raw]
            if any(str(c).strip() for c in cells):
                rows.append(cells)
        wb.close()
        return rows

    # Delimited text: csv / tsv / txt. Sniff a delimiter, fall back to comma.
    with open(path, encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        delimiter = ","
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except Exception:
            if lower.endswith(".tsv"):
                delimiter = "\t"
        reader = csv.reader(fh, delimiter=delimiter)
        for raw in reader:
            cells = ["" if c is None else c for c in raw]
            if any(str(c).strip() for c in cells):
                rows.append(cells)
    return rows


def parse_redirects_workbook(path: str) -> list[dict] | dict:
    """Read a redirects file and return old->new URL pairs.

    Returns a list of ``{"old_url": str, "new_url": str}`` on success, or an
    ``{"ok": False, "error": "..."}`` dict when the file is unreadable, empty,
    or contains no valid rows.
    """
    try:
        rows = _read_rows(path)
    except Exception as exc:  # unreadable file / missing sheet
        return {"ok": False, "error": f"Failed to read worksheet: {exc}"}

    if not rows:
        return {"ok": False, "error": "File is empty"}

    first = [str(c).lower().strip() for c in (rows[0] or [])]
    is_header = any(REDIRECT_HEADER_RE.search(c) for c in first)
    data_rows = rows[1:] if is_header else rows

    data: list[dict] = []
    for r in data_rows:
        old_url = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
        new_url = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        if old_url:
            data.append({"old_url": old_url, "new_url": new_url})

    if not data:
        return {"ok": False, "error": "The file contains no valid URLs"}
    return data


# ── clusters ─────────────────────────────────────────────────────────────────
def _cluster_fields(cluster: dict, index: int) -> tuple[Any, str, Any, list]:
    """Extract (id, name, count, keywords) from a cluster, tolerant of shape."""
    keywords = cluster.get("keywords") or []
    if not isinstance(keywords, list):
        keywords = list(keywords)
    name = cluster.get("label", cluster.get("name", ""))
    cid = cluster.get("id", index)
    count = cluster.get("count", len(keywords))
    return cid, "" if name is None else str(name), count, keywords


def safe_sheet_name(name: Any) -> str:
    """Excel-safe sheet name: <=30 chars, forbidden characters replaced with _."""
    cleaned = _FORBIDDEN_SHEET_CHARS.sub("_", str(name)[:30])
    return cleaned or "Sheet"


def build_clusters_csv(clusters: list[dict]) -> str:
    """Build a BOM-prefixed CSV string from clusters. Pure and testable."""

    def safe(v: Any) -> str:
        return '"' + str(v).replace('"', '""') + '"'

    rows = ["﻿Keyword,Cluster ID,Cluster name,Cluster size"]
    for index, cluster in enumerate(clusters, start=1):
        cid, name, count, keywords = _cluster_fields(cluster, index)
        for kw in keywords:
            rows.append(f"{safe(kw)},{cid},{safe(name)},{count}")
    return "\n".join(rows)


def write_clusters_csv(clusters: list[dict], path: str) -> dict | None:
    """Write clusters to a CSV file (UTF-8, BOM included in the content)."""
    try:
        with open(path, "w", encoding="utf-8", newline="") as fh:
            fh.write(build_clusters_csv(clusters))
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return None


def write_clusters_xlsx(clusters: list[dict], path: str) -> dict | None:
    """Write clusters to an xlsx file: one sheet per cluster plus an ``_All`` summary."""
    if Workbook is None:
        return {"ok": False, "error": "openpyxl is not available"}
    try:
        wb = Workbook()
        # Workbook() ships with one default sheet; remove it to build cleanly.
        wb.remove(wb.active)

        used: set[str] = set()
        for index, cluster in enumerate(clusters, start=1):
            _cid, name, count, keywords = _cluster_fields(cluster, index)
            title = safe_sheet_name(name)
            # Excel sheet titles must be unique; disambiguate if needed.
            base = title
            suffix = 2
            while title.lower() in used:
                tail = f"_{suffix}"
                title = base[: 30 - len(tail)] + tail
                suffix += 1
            used.add(title.lower())

            ws = wb.create_sheet(title=title)
            ws.append(["Keyword", "Cluster", "Cluster size"])
            for kw in keywords:
                ws.append([kw, name, count])

        summary = wb.create_sheet(title="_All")
        summary.append(["Keyword", "Cluster ID", "Cluster name"])
        for index, cluster in enumerate(clusters, start=1):
            cid, name, _count, keywords = _cluster_fields(cluster, index)
            for kw in keywords:
                summary.append([kw, cid, name])

        wb.save(path)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return None


# ── smoke test (no network / no files needed) ────────────────────────────────
if __name__ == "__main__":
    sample = [
        {"label": "pumps", "keywords": ["buy a pump", 'pump "CNS"']},
        {"id": 7, "label": "filters", "count": 1, "keywords": ["water filter"]},
    ]

    csv_text = build_clusters_csv(sample)
    lines = csv_text.split("\n")
    assert lines[0].startswith("﻿"), "CSV must start with a BOM"
    assert lines[0].endswith("Cluster size"), "unexpected header"
    # Quotes inside a keyword must be doubled.
    assert '"pump ""CNS"""' in csv_text, "quote escaping broken"
    # Derived id for the first cluster is its 1-based index.
    assert lines[1].split(",")[1] == "1", "derived id should be 1"
    # Explicit id is honoured.
    assert ",7," in csv_text, "explicit id not honoured"

    assert safe_sheet_name("a/b:c*d?e[f]") == "a_b_c_d_e_f_"
    assert safe_sheet_name("x" * 40) == "x" * 30
    assert safe_sheet_name("") == "Sheet"

    assert REDIRECT_HEADER_RE.search("old url")
    assert REDIRECT_HEADER_RE.search("откуда")
    assert not REDIRECT_HEADER_RE.search("https://example.com")

    print("excel.py self-check passed")
