"""Excel report with Summary, Findings, Pages, and Technologies worksheets.

The chart is built with Excel primitives through openpyxl rather than embedded as
a static matplotlib image. It therefore remains editable and tied to worksheet
data, while the generated workbook avoids an unnecessary plotting dependency.
"""

from __future__ import annotations

import pathlib
from typing import Any

_HEAD = {"critical": "C00000", "warning": "BF8F00", "notice": "808080"}


def _style_header(ws, row: int = 1) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F3864")
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # Use a string coordinate rather than ``ws.cell()``. Accessing a cell here
    # materializes a row, so the next ``append`` would skip one and leave a blank
    # row immediately below the header.
    ws.freeze_panes = f"A{row + 1}"


def _autofit(ws, limits: dict[int, int] | None = None) -> None:
    """Fit columns to content while capping widths for readable long-URL tables."""
    from openpyxl.utils import get_column_letter

    limits = limits or {}
    for idx, column in enumerate(ws.columns, start=1):
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(idx)].width = min(
            max(longest + 2, 10), limits.get(idx, 60)
        )


def write(document: dict[str, Any], path: pathlib.Path) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    summary = document.get("summary") or {}
    by_sev = summary.get("findings_by_severity") or {}

    # -- Summary -------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"SEO Audit: {document.get('domain', '')}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = document.get("url", "")
    ws["A3"] = f"Generated: {document.get('generated_at', '')}"

    rows = [
        ("Pages checked", summary.get("pages_checked", 0)),
        ("Total findings", summary.get("findings_total", 0)),
        ("Critical findings", by_sev.get("critical", 0)),
        ("Warnings", by_sev.get("warning", 0)),
        ("Notices", by_sev.get("notice", 0)),
        ("Checks completed", len(summary.get("tools_run") or [])),
        ("Checks unavailable", len(summary.get("tools_failed") or [])),
    ]
    ws["A5"], ws["B5"] = "Metric", "Value"
    _style_header(ws, 5)
    for i, (name, value) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=value)

    # The three severity bars expose the issue distribution at a glance.
    chart = BarChart()
    chart.title = "Findings by Severity"
    chart.y_axis.title = "Count"
    chart.add_data(Reference(ws, min_col=2, min_row=8, max_row=10), titles_from_data=False)
    chart.set_categories(Reference(ws, min_col=1, min_row=8, max_row=10))
    chart.legend = None
    chart.height, chart.width = 7, 12
    ws.add_chart(chart, "D5")

    failed = summary.get("tools_failed") or []
    if failed:
        start = 6 + len(rows) + 1
        ws.cell(row=start, column=1, value="Unavailable checks -- evidence is absent from report")
        ws.cell(row=start, column=1).font = Font(bold=True, color="C00000")
        for i, item in enumerate(failed, start=start + 1):
            ws.cell(row=i, column=1, value=item.get("tool"))
            ws.cell(row=i, column=2, value=item.get("error"))
    note = summary.get("severity_note")
    if note:
        ws.cell(row=6 + len(rows) + len(failed) + 3, column=1, value=note).font = Font(
            italic=True, size=9, color="808080"
        )
    _autofit(ws, {2: 40})

    # -- Findings ------------------------------------------------------------
    ws = wb.create_sheet("Findings")
    ws.append(["Severity", "Source", "URL", "Finding"])
    _style_header(ws)
    from seohead.reports import SEVERITY_TITLES

    for finding in document.get("findings") or []:
        ws.append(
            [
                SEVERITY_TITLES.get(finding.get("severity"), finding.get("severity")),
                finding.get("source", ""),
                finding.get("url", ""),
                finding.get("text", ""),
            ]
        )
        colour = _HEAD.get(finding.get("severity"))
        if colour:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color=colour)
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:D{ws.max_row}"
    _autofit(ws, {4: 100})

    # -- Pages ---------------------------------------------------------------
    pages = document.get("pages") or []
    ws = wb.create_sheet("Pages")
    columns = [
        "url",
        "status",
        "title",
        "title_length",
        "h1",
        "canonical",
        "words",
        "schema_types",
        "schema_errors",
        "social_missing",
    ]
    titles = [
        "URL",
        "Status",
        "Title",
        "Title Length",
        "H1",
        "Canonical",
        "Words",
        "Schema Types",
        "Schema Errors",
        "Missing Social Tags",
    ]
    ws.append(titles)
    _style_header(ws)
    for page in pages:
        ws.append([page.get(c, "") for c in columns])
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    _autofit(ws, {1: 70, 3: 60, 5: 40})

    # -- Technologies and infrastructure ------------------------------------
    ws = wb.create_sheet("Technologies")
    ws.append(["Category", "Detected Technology", "Evidence"])
    _style_header(ws)
    tech = (document.get("site") or {}).get("tech_detect") or {}
    for item in tech.get("technologies") or []:
        ws.append([item.get("category", ""), item.get("name", ""), item.get("evidence", "")])
    registration = ((document.get("site") or {}).get("domain_profile") or {}).get(
        "registration"
    ) or {}
    if registration:
        ws.append([])
        ws.append(["domain", "registrar", registration.get("registrar", "")])
        ws.append(["domain", "created", registration.get("created", "")])
        ws.append(["domain", "expires", registration.get("expires", "")])
        ws.append(["domain", "age in years", registration.get("age_years", "")])
    _autofit(ws, {3: 70})

    wb.save(path)
